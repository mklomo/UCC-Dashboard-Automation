import sqlite3

import openpyxl

# Week to UPDATE (INTEGER)
WEEK_TO_UPDATE = 10

# Connection to DB 
CONNECTION = None 

# Cursor to execute Commands
CURSOR = None

# Provide the Relative Path to the Excel File 
EXCEL_FILE = "data\master_data.xlsx"


# Create a work book object
WB = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

# Select the Feedback Sheet from the workbook
FEEDBACK_WORKSHEET = WB['Feedback']

# PART 3 - UPDATING RECORDS IN THE ASSIGNMENT TABLE
def db_connection(path_to_db:str):
    """This function creates a connection to the DB using the path specified
    

    Args:
        path_to_db (str): Specify the location of the DB
    """
    global CONNECTION, CURSOR
    try:

        # Connect to a database or Create a DB
        CONNECTION = sqlite3.connect(path_to_db)
        
        # Get the cursor ready 
        CURSOR = CONNECTION.cursor()

        # Check the Cursor Connection
        print("Successfully Connected to DB")
        
    except sqlite3.Error as error:
        print("Failed to connect to DB", error)
    
    
    
def update_week_records(week_number: int, path_to_excel:str):
    """
        This function updates entries the records (weekly_score, student_status, comments)
        for a student_name in a specified week
        week_number: this is an integer from 1 - 10
    """
    # Refer to global variables
    global CONNECTION, CURSOR
    
    # Get the record list for the week from the excel sheet
    
    # Get the Student Names
    student_names_col = FEEDBACK_WORKSHEET.iter_cols(min_row=3, max_row=55, min_col=1, max_col=1)
    
    # Get the student names
    student_names = []
    
    for names_tuple in student_names_col:
        for name in names_tuple:
            student_names.append(name.value)
    
    # Weekly Column list for weeks 1 - 10
    column_numbers = [11, 20, 29, 38, 47, 56, 65, 74, 83, 92]
    
    # Select the total_score, comment and status column for each week
    start_col_num = column_numbers[week_number-1]
    end_col_num = start_col_num+2
    assignment_columns = FEEDBACK_WORKSHEET.iter_rows(min_row=3, max_row=55, min_col=start_col_num, max_col=end_col_num)


    # Fields of interest in the assignment table are - week, weekly_score, student_status, comments

    # Get a list of of all week  data
    week_list = [f"week {week_number}"] * len(student_names)
    week_score = []
    week_comment = []
    week_status = []

    for score, comment, status in assignment_columns:
        # Assign the values to the respective list
        week_score.append(score.value)
        week_comment.append(comment.value)
        week_status.append(status.value)
    

    new_week_record_list = []

    for i in range(len(student_names)):
        # Create the record
        record = (week_score[i], week_comment[i], week_status[i], week_list[i],student_names[i])
        # Append the record to the student master data
        new_week_record_list.append(record)
      
    
    try:        
        assignment_update_query = """
                            UPDATE assignments set weekly_score = ?,
                                                          comments = ?,
                                                      student_status = ?                                                 
                                                           where week = ? AND
                                                         student_name = ?   
                                """
    
        # Now, execute the query
        CURSOR.executemany(assignment_update_query, new_week_record_list)
        
        # Commit the Work
        CONNECTION.commit()

        
        
    except sqlite3.Error as error:
        print("Failed to update multiple records of sqlite table", error)
        
    else:
        print("DB Has been successfully updated 👍")    
       
                                         
                                         
                                         
def main():
    
    # Provide the DB Path
    db_path = r"database\ucc.db"
    
    # Connect to the DB 
    db_connection(path_to_db=db_path)
    
    with CONNECTION:
        # Update record for week
        update_week_records(path_to_excel=EXCEL_FILE, week_number=WEEK_TO_UPDATE)
        
     
if __name__ == '__main__':
    main()    
    