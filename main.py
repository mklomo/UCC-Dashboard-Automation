import sqlite3

import openpyxl


# Connection to DB 
CONNECTION = None 

# Cursor to execute Commands
CURSOR = None

# Provide the Relative Path to the Excel File 
EXCEL_FILE = "data\master_data.xlsx"

# List of Students in the cohort 
STUDENT_NAMES = None

# Create a work book object
WB = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

# Select the Feedback Sheet from the workbook
FEEDBACK_WORKSHEET = WB['Feedback']


def db_connection(path_to_db:str):
    """This function creates a connection to the DB using the path specified
    

    Args:
        path_to_db (str): Specify the location of the DB
    """
    global CONNECTION, CURSOR
    try:

        # Connect to a database or Create a DB
        CONNECTION = sqlite3.connect(path_to_db)
        
        # Get the cursor ready 
        CURSOR = CONNECTION.cursor()

        # Check the Cursor Connection
        print("Successfully Connected to DB")
        
    except sqlite3.Error as error:
        print("Failed to connect to DB", error)
    
    


    # Inserting records into the students table - row-wise insert
def insert_student_records():
    """
        This function inserts records into the students Master Table
        
        db_connection: a db connection object
    """
    # Reference the global variable
    global STUDENT_NAMES

    # Accessing the student details in the feedback Sheet - student_name, gender, track, facilitator_name
    STUDENT_NAMES = []
    student_gender = []
    student_track = []
    student_fac_name = []

    # Iterate from row 3 to row 55 in colimns 1 - 4
    student_column = FEEDBACK_WORKSHEET.iter_rows(min_row=3, max_row=55, min_col=1, max_col=4)

    
    for name, gender, track, fac_name in student_column:
        # Append the value to each list
        STUDENT_NAMES.append(name.value)
        student_gender.append(gender.value)
        student_track.append(track.value)
        student_fac_name.append(fac_name.value)


    # Preparing data for DB insertion - by creating a tuple for each student record 
       
    student_master_data  = [] 
    
    for i in range(len(STUDENT_NAMES)):
        # Create the record
        record = (STUDENT_NAMES[i], student_gender[i], student_track[i], student_fac_name[i])
        # Append the redord to the student master data
        student_master_data.append(record)


    # Insert the student Master Data
    query = """
                        INSERT INTO students 
                        VALUES (?,?,?,?)

        """
    
        
    CURSOR.executemany(query, student_master_data)
    
    # Commit the Data
    CONNECTION.commit()
    
   
    # Tell User Record Insertion is Successful
    print("Records in student table inserted successfully! 👍")
        


# Assignment Table
def insert_assignments_data():
    global STUDENT_NAMES
    
    # Weekly Column list for weeks 1 - 10 from yhe excel workbook
    column_numbers = [11, 20, 29, 38, 47, 56, 65, 74, 83, 92]
        
    # Assignment Master Data
    assignment_master_data = []

    for column in range(len(column_numbers)):
        # Select the total_score, comment and status column for each week
        start_col_num = column_numbers[column]
        end_col_num = start_col_num+2
        assignment_columns = FEEDBACK_WORKSHEET.iter_rows(min_row=3, max_row=55, min_col=start_col_num, max_col=end_col_num)


        # Fields of interest in the assignment table are - week, weekly_score, student_status, comments

        # Get a list of of all week  data
        week_list = [f'week {column+1}'] * len(STUDENT_NAMES)
        week_score = []
        week_comment = []
        week_status = []


        for score, comment, status in assignment_columns:
            # Assign the values to the respective list
            week_score.append(score.value)
            week_comment.append(comment.value)
            week_status.append(status.value)
    

        # Create the Week Record  -- student_name -- week -- weekly_score -- student_status -- comments

        for i in range(len(STUDENT_NAMES)):
            # Create the record
            record = (week_list[i], week_score[i], week_comment[i], week_status[i], STUDENT_NAMES[i])
            # Append the record to the student master data
            assignment_master_data.append(record)
        
    
            
    query = """
                    INSERT INTO assignments 
                    VALUES (?,?,?,?,?)

        """

        
    # Execute the command
    CURSOR.executemany(query, assignment_master_data)
    
    # Commit the Transaction
    CONNECTION.commit()

    print("Assignment Data Inserted Successfully 👍")

   

##########################################################################################################################    


def main():


    # Create a DB Connection
    db_path = r'database\ucc.db'

    # Connection to the Database
    db_connection(path_to_db=db_path)

    
    with CONNECTION:
        # Creating Tables in the UCC Database using the cursor

        students_tb_query = """CREATE TABLE IF NOT EXISTS students (
                                                            student_name TEXT PRIMARY KEY,
                                                            gender TEXT,
                                                            track TEXT,
                                                            facilitator_name TEXT
                                                        ) 
                        """


        assignments_tb_query = """CREATE TABLE IF NOT EXISTS assignments (
                                                        week TEXT,
                                                        weekly_score INTEGER,
                                                        student_status TEXT,
                                                        comments TEXT,
                                                        student_name TEXT,
                                                        PRIMARY KEY (week, student_name),
                                                        FOREIGN KEY (student_name) REFERENCES students (student_name) 
                                                        ) 
                                """


                    
        try:
            CURSOR.execute(students_tb_query)
            
            # Commit the Result 
            CONNECTION.commit()


        except sqlite3.Error as error:
            print(error)
            
            
        try:    
            CURSOR.execute(assignments_tb_query)  
            
            # Commit the Result 
            CONNECTION.commit()  

        except sqlite3.Error as error:
            print("Unable to create the assignments table", error)
            

        else:
            # Insert records into the student table in the DB 
            insert_student_records()


            # Insert records into the assignments table in the DB 
            insert_assignments_data()


            # Print a Congratulatory Message 
            print("Congratulations, ALL Records have been successfully loaded into the DB 🥳🥳🥳🥳")
        
   

if __name__ == '__main__':
    main()
